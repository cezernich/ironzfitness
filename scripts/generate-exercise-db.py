#!/usr/bin/env python3
"""
Generate exercise-data.js from IronZ_Exercise_Library_Expanded.xlsx.

Reads all 4 sheets (Strength, Circuit & Bodyweight, Hyrox Stations,
Sport-Specific Strength), normalizes every row into the schema defined
in cowork-handoff/EXERCISE_DB_SPEC.md, and emits exercise-data.js with
window.EXERCISE_DB as a flat array.

Idempotent — re-running with the same spreadsheet produces identical
output (deterministic id generation + sorted output).

Usage:
    python3 scripts/generate-exercise-db.py
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install --user openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "IronZ_Exercise_Library_Expanded.xlsx"
OUT_PATH  = REPO_ROOT / "exercise-data.js"


# ── Normalization tables ──────────────────────────────────────────────────────

def kebab(s: str) -> str:
    """kebab-case from any string. Strips diacritics and punctuation."""
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


# Equipment string → canonical equipmentNeeded tokens. The xlsx uses
# free-form English ("Barbell & Rack, Bench"); we split on commas, then
# map each fragment to one or more canonical tokens. Slash-separated
# alternatives ("Cable/Band") expand into separate tokens too.
EQUIPMENT_FRAGMENT_MAP = {
    "barbell & rack":           ["barbell-rack"],
    "barbell rack":             ["barbell-rack"],
    "barbell":                  ["barbell-rack"],   # bare "barbell" implies rack in this app
    "rack":                     ["barbell-rack"],
    "dumbbells":                ["dumbbells"],
    "dumbbell":                 ["dumbbells"],
    "db":                       ["dumbbells"],
    "kettlebell":               ["kettlebell"],
    "kb":                       ["kettlebell"],
    "bench":                    ["bench"],
    "pull-up bar":              ["pull-up-bar"],
    "pull up bar":              ["pull-up-bar"],
    "cable":                    ["cable-machine"],
    "cables":                   ["cable-machine"],
    "cables / machines":        ["cable-machine"],
    "cables/machines":          ["cable-machine"],
    "machine":                  ["cable-machine"],
    "machines":                 ["cable-machine"],
    "cable crossover":          ["cable-machine"],
    "functional trainer":       ["functional-trainer"],
    "lat pulldown":             ["lat-pulldown"],
    "seated row":               ["seated-row"],
    "leg press":                ["leg-press"],
    "leg curl":                 ["leg-curl"],
    "leg extension":            ["leg-extension"],
    "smith":                    ["smith-machine"],
    "smith machine":            ["smith-machine"],
    "ghd":                      ["ghd"],
    "ab wheel":                 ["ab-wheel"],
    "band":                     ["band"],
    "bands":                    ["band"],
    "jump rope":                ["jump-rope"],
    "med ball":                 ["med-ball"],
    "rowing machine":           ["rowing-machine"],
    "ski erg":                  ["ski-erg"],
    "skierg":                   ["ski-erg"],
    "sled":                     ["sled"],
    "sandbag":                  ["sandbag"],
    "trap bar":                 ["trap-bar"],
    "weight plate":             ["weight-plate"],
    "plate":                    ["weight-plate"],
    "ez bar":                   ["barbell-rack"],   # treat EZ bar as barbell-class
    "landmine":                 ["barbell-rack"],
    "hip abductor/adductor":    ["hip-abductor-adductor"],
    "chest press machine":      ["chest-press-machine"],
    "chest fly machine":        ["chest-fly-machine"],
    "shoulder press machine":   ["shoulder-press-machine"],
    "bodyweight":               [],   # canBeBodyweight handles this; no equipment token
}


def normalize_equipment(raw: str | None) -> list[str]:
    """Map a free-form equipment string to a deduped list of canonical tokens."""
    if not raw:
        return []
    out: list[str] = []
    # Split on commas; within each fragment also split on slash for "A/B" alternatives
    for fragment in str(raw).split(","):
        fragment = fragment.strip()
        if not fragment:
            continue
        # Try direct match first
        key = fragment.lower()
        if key in EQUIPMENT_FRAGMENT_MAP:
            for tok in EQUIPMENT_FRAGMENT_MAP[key]:
                if tok and tok not in out:
                    out.append(tok)
            continue
        # Otherwise split slash-alternatives
        for alt in fragment.split("/"):
            alt_key = alt.strip().lower()
            if alt_key in EQUIPMENT_FRAGMENT_MAP:
                for tok in EQUIPMENT_FRAGMENT_MAP[alt_key]:
                    if tok and tok not in out:
                        out.append(tok)
            else:
                # Unknown — slug it as a fallback so it doesn't silently disappear
                slug = kebab(alt)
                if slug and slug not in out:
                    out.append(slug)
    return out


# Specific muscle goal → canonical sub-target token. Composite values
# ("Quads + glutes, Adductors") use the first comma fragment as the primary.
SPECIFIC_GOAL_MAP = {
    "quads + glutes":                          "quads-glutes",
    "quads emphasis":                          "quads-emphasis",
    "quads (knee extension)":                  "quads-emphasis",
    "posterior chain (glutes + hamstrings)":   "posterior-chain",
    "glutes (hip extension)":                  "glutes-hip-extension",
    "hamstrings (knee flexion)":               "hamstrings-knee-flexion",
    "erectors / lower back":                   "erectors-lower-back",
    "upper chest":                             "upper-chest",
    "lower chest":                             "lower-chest",
    "chest stretch / pec isolation":           "chest-isolation",
    "triceps":                                 "triceps",
    "overhead strength":                       "overhead-strength",
    "side delts":                              "side-delts",
    "front delts":                             "front-delts",
    "rear delts + scapular stability":         "rear-delts-scapular",
    "mid-back (rhomboids) + lats":             "mid-back-lats",
    "lats (vertical pull)":                    "lats-vertical",
    "core stability":                          "core-stability",
    "anti-rotation":                           "anti-rotation",
    "obliques":                                "obliques",
    "lower abs / hip flexors":                 "lower-abs-hip-flexors",
    "rectus abdominis":                        "rectus-abdominis",
    "transverse abs":                          "core-stability",
    "biceps":                                  "biceps",
    "biceps + brachialis":                     "biceps-brachialis",
    "calves":                                  "calves",
    "glute medius":                            "glute-medius",
    "adductors":                               "adductors",
    "hip flexors":                             "lower-abs-hip-flexors",
    "rotator cuff":                            "rear-delts-scapular",
    "rear delts":                              "rear-delts-scapular",
    "front delts":                             "front-delts",
    "side delts":                              "side-delts",
    "lower abs":                               "lower-abs-hip-flexors",
}


def normalize_specific_goal(raw: str | None) -> str | None:
    if not raw:
        return None
    primary = str(raw).split(",")[0].strip().lower()
    if primary in SPECIFIC_GOAL_MAP:
        return SPECIFIC_GOAL_MAP[primary]
    # Fallback to slug — keeps unknown goals visible rather than dropping them
    return kebab(primary)


def normalize_pattern(raw: str | None) -> str | None:
    if not raw:
        return None
    return kebab(raw)


def normalize_muscle_category(raw: str | None) -> list[str]:
    """Comma-separated list of muscle categories → kebab-case tokens, deduped."""
    if not raw:
        return []
    out: list[str] = []
    for m in str(raw).split(","):
        slug = kebab(m)
        if slug and slug not in out:
            out.append(slug)
    return out


def yes_to_bool(raw) -> bool:
    if raw is None:
        return False
    return str(raw).strip().lower() in ("yes", "true", "y", "1")


def normalize_modality(raw: str | None) -> str | None:
    if not raw:
        return None
    m = str(raw).strip().lower()
    return {
        "bodyweight": "bodyweight",
        "barbell":    "barbell",
        "kettlebell": "kettlebell",
        "db/kb":      "dumbbell-kettlebell",
        "med ball":   "med-ball",
        "cardio":     "cardio",
        "machine":    "machine",
        "ghd":        "ghd",
        "jump rope":  "jump-rope",
        "band/bw":    "bodyweight-band",
    }.get(m, kebab(m))


def normalize_common_in(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [kebab(x) for x in str(raw).split(",") if x.strip()]


def normalize_sport(raw: str | None) -> str | None:
    if not raw:
        return None
    return {"swim": "swim", "cycling": "cycling", "running": "running"}.get(
        str(raw).strip().lower(), kebab(raw))


# ── Row → exercise object ─────────────────────────────────────────────────────

def make_id(name: str, used_ids: set[str]) -> str:
    """kebab-case unique id. Collisions get -2, -3, ... suffix (alphabetical)."""
    base = kebab(name)
    if base not in used_ids:
        used_ids.add(base)
        return base
    n = 2
    while f"{base}-{n}" in used_ids:
        n += 1
    final = f"{base}-{n}"
    used_ids.add(final)
    return final


def row_strength(row: list, used_ids: set[str]) -> dict | None:
    pattern, name, tier, equip, primary, cat, goal, weights, bw, needed = row[:10]
    if not name:
        return None
    return {
        "id": make_id(str(name), used_ids),
        "name": str(name).strip(),
        "sheet": "strength",
        "pattern": normalize_pattern(pattern),
        "tier": (str(tier).strip().lower() if tier else None),
        "equipmentTags": normalize_equipment(equip),
        "primaryMuscles": (str(primary).strip() if primary else None),
        "muscleCategory": normalize_muscle_category(cat),
        "specificGoal": normalize_specific_goal(goal),
        "usesWeights": yes_to_bool(weights),
        "canBeBodyweight": yes_to_bool(bw),
        "equipmentNeeded": normalize_equipment(needed),
        "modality": None,
        "commonIn": None,
        "sport": None,
        "purpose": None,
        "isHyroxStation": False,
        "hyroxOrder": None,
        "defaultDistance": None,
        "defaultWeight": None,
    }


def row_circuit(row: list, used_ids: set[str]) -> dict | None:
    category, name, modality, common, cat, goal, weights, bw, needed = row[:9]
    if not name:
        return None
    return {
        "id": make_id(str(name), used_ids),
        "name": str(name).strip(),
        "sheet": "circuit",
        "pattern": None,
        "tier": None,
        "equipmentTags": [],
        "primaryMuscles": None,
        "muscleCategory": normalize_muscle_category(cat),
        "specificGoal": normalize_specific_goal(goal),
        "usesWeights": yes_to_bool(weights),
        "canBeBodyweight": yes_to_bool(bw),
        "equipmentNeeded": normalize_equipment(needed),
        "modality": normalize_modality(modality),
        "commonIn": normalize_common_in(common),
        "sport": None,
        "purpose": None,
        "isHyroxStation": False,
        "hyroxOrder": None,
        "defaultDistance": None,
        "defaultWeight": None,
        # Keep the spreadsheet's own category column (Upper Body BW, Core, ...)
        "circuitCategory": (str(category).strip() if category else None),
    }


def row_hyrox(row: list, used_ids: set[str]) -> dict | None:
    order, name, dist, weight, _icon = row[:5]
    if not name:
        return None
    return {
        "id": make_id(str(name), used_ids),
        "name": str(name).strip(),
        "sheet": "hyrox",
        "pattern": None,
        "tier": None,
        "equipmentTags": [],
        "primaryMuscles": None,
        "muscleCategory": [],
        "specificGoal": None,
        "usesWeights": bool(weight and str(weight).strip() not in ("-", "")),
        "canBeBodyweight": False,
        "equipmentNeeded": [],
        "modality": None,
        "commonIn": ["hyrox"],
        "sport": None,
        "purpose": None,
        "isHyroxStation": True,
        "hyroxOrder": int(order) if order else None,
        "defaultDistance": (str(dist).strip() if dist else None),
        "defaultWeight": (str(weight).strip() if weight and str(weight).strip() != "-" else None),
    }


def row_sport_specific(row: list, used_ids: set[str]) -> dict | None:
    sport, name, purpose, cat, goal, weights, bw, needed = row[:8]
    if not name:
        return None
    return {
        "id": make_id(str(name), used_ids),
        "name": str(name).strip(),
        "sheet": "sport-specific",
        "pattern": None,
        "tier": None,
        "equipmentTags": [],
        "primaryMuscles": None,
        "muscleCategory": normalize_muscle_category(cat),
        "specificGoal": normalize_specific_goal(goal),
        "usesWeights": yes_to_bool(weights),
        "canBeBodyweight": yes_to_bool(bw),
        "equipmentNeeded": normalize_equipment(needed),
        "modality": None,
        "commonIn": None,
        "sport": normalize_sport(sport),
        "purpose": (str(purpose).strip() if purpose else None),
        "isHyroxStation": False,
        "hyroxOrder": None,
        "defaultDistance": None,
        "defaultWeight": None,
    }


# ── Driver ────────────────────────────────────────────────────────────────────

def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: spreadsheet not found at {XLSX_PATH}", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX_PATH, data_only=True)
    out: list[dict] = []
    used_ids: set[str] = set()

    sheet_handlers = {
        "Strength":                ("strength",       row_strength,       10),
        "Circuit & Bodyweight":    ("circuit",        row_circuit,         9),
        "Hyrox Stations":          ("hyrox",          row_hyrox,           5),
        "Sport-Specific Strength": ("sport-specific", row_sport_specific,  8),
    }

    for sheet_name, (label, handler, expected_cols) in sheet_handlers.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARN: missing sheet {sheet_name!r}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        added = 0
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, expected_cols + 1)]
            obj = handler(row, used_ids)
            if obj:
                out.append(obj)
                added += 1
        print(f"  {sheet_name}: {added} exercises", file=sys.stderr)

    # Sort by sheet then id for deterministic output
    sheet_order = {"strength": 0, "circuit": 1, "hyrox": 2, "sport-specific": 3}
    out.sort(key=lambda x: (sheet_order.get(x["sheet"], 99), x["id"]))

    print(f"\nTotal: {len(out)} exercises", file=sys.stderr)

    # Emit as window.EXERCISE_DB. Pretty-printed for diffability.
    json_blob = json.dumps(out, indent=2, ensure_ascii=False)
    contents = (
        "// exercise-data.js — GENERATED FROM IronZ_Exercise_Library_Expanded.xlsx\n"
        "//\n"
        "// DO NOT EDIT BY HAND. Edit the spreadsheet and run:\n"
        "//   python3 scripts/generate-exercise-db.py\n"
        "//\n"
        "// Schema: cowork-handoff/EXERCISE_DB_SPEC.md\n"
        "// Source of truth: IronZ_Exercise_Library_Expanded.xlsx (4 sheets)\n"
        "\n"
        "(function () {\n"
        '  "use strict";\n'
        f"  window.EXERCISE_DB = {json_blob};\n"
        "})();\n"
    )
    OUT_PATH.write_text(contents, encoding="utf-8")
    print(f"Wrote {OUT_PATH.relative_to(REPO_ROOT)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
